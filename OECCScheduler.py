""" Represents the entire Scheduler window including the button functionality and graphics scene """
""" Data keys used in graphics objects: 
        0 = (segment) block id placement; (procedure) beginning block, 
        1 = (procedure) length of block, 
        2 = (procedure) block procedure identifier, 
        3 = (3-type text, 2-date text, 1-name text) text object identifier, 
        4 = block segment identifier """

import os
import numpy as np
import openpyxl as op
from datetime import date

from BlockSegment import BlockSegment
from GraphicsScene import GraphicsScene
from GraphicsRectItem import GraphicsRectItem
from ScheduleColors import BlockColors
from Guide import HelpGuide
from CheckableComboBox import CheckableComboBox
import ProcedureDicts

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QBrush, QPainter, QPen, QColor, QIcon, QFont, QPixmap
from PyQt6.QtWidgets import (
    QComboBox,
    QFileDialog,
    QDoubleSpinBox,
    QGraphicsItem,
    QGraphicsRectItem,
    QGraphicsTextItem,
    QGraphicsView,
    QHBoxLayout,
    QPushButton,
    QLineEdit,
    QVBoxLayout,
    QWidget,
    QLabel,
    QGroupBox,
    QFormLayout,
)       

class Window(QWidget):

    canvasWidth = 100
    canvasHeight = 100
    schedule = []
    timeBoxWidth = 80
    blockSize = 40
    blockWidth = 100
    timeWidth = 50
    typeWidth = 70
    numSubBlocks = 4
    currentCaseCount = 0
    blockType = "Regular"
    blockName = "Patient"
    blockModName = "Patient"
    blockTimeCustom = 1
    customBlockLength = 1.00
    blockTimes = ProcedureDicts.blockTimes
    blockColors = ProcedureDicts.blockColors
    blockEndMinutes = {
        0 : { 0 : "03", 1 : "07", 2 : "11", 3 : "14",},
        15 : { 0 : "18", 1 : "22", 2 : "26", 3 : "39",},
        30 : { 0 : "33", 1 : "37", 2 : "41", 3 : "44",},
        45 : { 0 : "48", 1 : "52", 2 : "56", 3 : "59",},
    }
    blockBeginMinutes = {
        0 : { 0 : "00", 1 : "04", 2 : "08", 3 : "12",},
        15 : { 0 : "15", 1 : "19", 2 : "23", 3 : "27",},
        30 : { 0 : "30", 1 : "34", 2 : "38", 3 : "42",},
        45 : { 0 : "45", 1 : "49", 2 : "53", 3 : "57",},
    }
    firstEmptyBlock = 0
    scaleAddition = 0

    guideWindow = None

    def __init__(self, saved=None, baseDir=""):
        super().__init__()

        # Defining scene and adding basic layout elements to scene
        numHours = 11
        numBlocks = self.numSubBlocks * numHours + 1
        startHour = 7
        startMinute = 30
        AMPMTag = "AM"
        segBoxWidth = self.blockWidth + self.timeWidth + self.typeWidth + 40
        topWhiteSpace = 24

        self.baseDir = baseDir
        self.canvasHeight = self.blockSize * numBlocks + topWhiteSpace
        self.canvasWidth = self.timeBoxWidth + segBoxWidth

            # Defining a scene rect, with it's origin at 0,-24 along with background & date time plus case count.
        self.setWindowTitle("Scheduler")
        self.scene = GraphicsScene(0, -1 * topWhiteSpace, self.canvasWidth, self.canvasHeight, self)

        self.scene.addRect(0, -1 * topWhiteSpace, self.canvasWidth, self.canvasHeight, brush = QBrush(Qt.GlobalColor.white))
        
        dateItem = self.scene.addText("Today's date: " + str(date.today()))
        dateItem.setPos(10, -1 * topWhiteSpace)
        self.caseCountItem = self.scene.addText("Cases: " + str(self.currentCaseCount))
        self.caseCountItem.setPos(180, -1 * topWhiteSpace)


            # Add time slots & corresponding boxes to graphics scene
        for i in range(0,numBlocks):
            if startHour < 10:
                textHour = "0" + str(startHour)
            else:
                textHour = str(startHour)
            
            if startMinute == 0:
                textMinute = "00"
            else:
                textMinute = str(startMinute)
            
            textItem = self.scene.addText( textHour + ":" + textMinute + " " + AMPMTag)
            textItem.setPos(0, self.blockSize * i)

            timeBox = QGraphicsRectItem(0, 0, self.timeBoxWidth, self.blockSize)
            timeBox.setBrush(QColor(255, 0, 0, 0))
            timeBox.setPen(QPen(Qt.GlobalColor.black))
            timeBox.setPos(0, self.blockSize * i)
            self.scene.addItem(timeBox)
            for j in range(self.numSubBlocks):
                blockSegBox = QGraphicsRectItem(0, 0, segBoxWidth, (self.blockSize / self.numSubBlocks))
                blockSegBox.setPos(80, (self.blockSize * i) + ((self.blockSize / self.numSubBlocks) * j))
                blockSegBox.setBrush(QColor(0, 0, 0, 0))
                blockSegBox.setPen(QPen(Qt.GlobalColor.gray))
                blockSegBox.setData(0, i * self.numSubBlocks + j)   # Set 0 to be the id of block segment
                blockSegBox.setData(4, 0)           # Set 4 to be identifier for a block segment
     
                #           UTILITY NUMBERING OF BOXES
                # blockSegBoxText = QGraphicsTextItem(str(i *4 + j), blockSegBox)   
                # blockSegBoxText.setPos(140, -4)
                # blockSegBoxText.setFont(QFont("Helvetica", 6))
                #           END UTIL
                
                self.scene.addItem(blockSegBox)
                self.scene.schedule.append(BlockSegment(i * self.numSubBlocks + j, (self.blockSize * i) + ((self.blockSize / self.numSubBlocks) * j),
                                                        textHour + ":" + self.blockBeginMinutes[startMinute][j],
                                                        textHour + ":" + self.blockEndMinutes[startMinute][j]))
            blockBox = QGraphicsRectItem(80, self.blockSize * i, segBoxWidth, self.blockSize)
            blockBox.setBrush(QColor(255, 0, 0, 0))
            blockBox.setPen(QPen(Qt.GlobalColor.black))
            self.scene.addItem(blockBox)

            startMinute += 15
            if startMinute >= 60:
                startMinute = 0
                startHour += 1
                if startHour >= 13:
                    startHour = 1
                elif startHour == 12:
                    if AMPMTag == "AM":
                        AMPMTag = "PM"
                    else:
                        AMPMTag = "AM"



            # Insert saved data. If no data, insert a break block at 12 pm
                # SAVED DATA FORMAT: (TYPE, NAME ,SEGMENT LENGTH, ALL TYPES)
        breakFlag = True
        skipCount = 0

        try:
            for savedBlock in saved:
                inputs = savedBlock.split(",")
                
                if int(inputs[0]) == 0:
                    skipCount += 1
                else:
                    breakFlag = False
                    blockTypes = []
                    if len(inputs[3]) < 1:
                        inputs[3] = "0"
                    for type in inputs[3].split("-"):
                        blockTypes.append(ProcedureDicts.procedureList[int(type)])
                    if len(inputs[1]) < 1:
                        inputs[1] = "Patient"
                    self.insertSaved(int(inputs[0]), skipCount, inputs[1], int(inputs[2]), blockTypes)
                    skipCount += int(inputs[2])
                
        except Exception as e:
            print("Could not insert saved procedures: ", e)

        if breakFlag:
            # Add breaktime if no saved data
            breakBlockLength = 2
            breakStart = 18 * self.numSubBlocks
            breakBlock = GraphicsRectItem(0, 0, self.blockWidth, self.blockSize * breakBlockLength, self.timeWidth, self.typeWidth,
                                          self.getTimeText(breakStart, breakStart + int(self.numSubBlocks * breakBlockLength) - 1), 
                                          "Break", ["Break"], QBrush(BlockColors.BREAK.value), int(self.numSubBlocks * breakBlockLength))
            breakBlock.setPos(self.timeBoxWidth, self.scene.schedule[breakStart].y)
            breakBlock.setData(0, breakStart)                                     # id of first block segment that break occupies
            breakBlock.setData(1, breakBlockLength * self.numSubBlocks)                         # number of segments that break occupies
            breakBlock.setData(2, ProcedureDicts.procedureIDs["Break"])          # identifier that this is a movable block
            
            self.scene.addItem(breakBlock)
            self.scene.procedures.append(breakBlock)
            breakBlock.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable)
            breakBlock.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable)
            self.fill(breakStart, self.numSubBlocks * breakBlockLength, True)
            


        # Define our layout & add functionality buttons
        vbox = QVBoxLayout()

            # INSERT LAYOUT BOX

        self.patientName = QLineEdit()
        self.patientName.setPlaceholderText("Patient Name")
        self.patientName.textChanged.connect(self.name_changed)

        insert = QPushButton("Insert")
        insert.setIcon(QIcon(os.path.join(self.baseDir, "icons", "insert.png")))
        insert.clicked.connect(self.insert)

        setBlock = QComboBox()
        setBlock.addItems(["Custom", "Break"] + ProcedureDicts.procedureList[3:(len(ProcedureDicts.procedureList) - 1)])
        setBlock.setCurrentIndex(2)
        setBlock.currentTextChanged.connect(self.text_changed)

        self.customLength = QDoubleSpinBox()
        self.customLength.setMinimum(0.25)
        self.customLength.setValue(1.00)
        self.customLength.setSingleStep(0.25)
        self.customLength.setEnabled(False)
        self.customLength.valueChanged.connect(self.changeCustomLength)
        self.customLengthLabel = QLabel("Length:")
        self.customLengthLabel.setBuddy(self.customLength)
        self.customLengthLabel.setEnabled(False)

        # comunes = ['Ameglia', 'Arcola', 'Bagnone', 'Bolano']
        # self.combo = CheckableComboBox()
        # self.combo.addItems(comunes)
        # #self.combo.toggleItem(2)
        # self.combo.currentTextChanged.connect(self.textChanged2)

        insertGroupBox = QGroupBox("Insert Procedure")
        insertGroupLayout = QFormLayout()
        insertGroupLayout.addRow(self.patientName)
        insertGroupLayout.addRow(QLabel("Procedure:"), setBlock)
        #insertGroupLayout.addRow(QLabel("Procedure:"), self.combo)
        insertGroupLayout.addRow(self.customLengthLabel, self.customLength)
        insertGroupLayout.addRow(insert)
        insertGroupLayout.setSizeConstraint(QFormLayout.SizeConstraint.SetFixedSize)
        insertGroupBox.setLayout(insertGroupLayout)
        vbox.addWidget(insertGroupBox)

            # END INSERT LAYOUT BOX
            # MODIFY LAYOUT BOX

        self.modifyPatientName = QLineEdit()
        self.modifyPatientName.setPlaceholderText("Patient Name")
        self.modifyPatientName.textChanged.connect(self.nameModified)

        self.deleteBlock = QPushButton("Delete")
        self.deleteBlock.setEnabled(False)
        self.deleteBlock.setIcon(QIcon(os.path.join(self.baseDir, "icons", "delete.png")))
        self.deleteBlock.clicked.connect(self.delete)

        self.changeName = QPushButton("Change Name")
        self.changeName.setEnabled(False)
        self.changeName.setIcon(QIcon(os.path.join(self.baseDir, "icons", "changeName.png")))
        self.changeName.clicked.connect(self.changePatientName)

        modifyGroupBox = QGroupBox("Modify Procedure")
        modifyGroupLayout = QFormLayout()
        modifyGroupLayout.addRow(self.modifyPatientName)
        modifyGroupLayout.addRow(self.changeName, self.deleteBlock)
        modifyGroupLayout.setSizeConstraint(QFormLayout.SizeConstraint.SetFixedSize)
        modifyGroupBox.setLayout(modifyGroupLayout)
        vbox.addWidget(modifyGroupBox)

            # END MODIFY LAYOUT BOX

        clear = QPushButton("Clear")
        clear.setIcon(QIcon(os.path.join(self.baseDir, "icons", "clear.png")))
        clear.clicked.connect(self.clear)
        vbox.addWidget(clear)

        openFile = QPushButton("Open")
        openFile.setIcon(QIcon(os.path.join(self.baseDir, "icons", "open.png")))
        openFile.clicked.connect(self.openFile)
        vbox.addWidget(openFile)

        # UTILITY BUTTON
        reset = QPushButton("Reset")
        reset.setIcon(QIcon(os.path.join(self.baseDir, "icons", "reset.png")))
        reset.clicked.connect(self.darkenFull)
        vbox.addWidget(reset)
        # UTIL END

        squish = QPushButton("Collapse")
        squish.setIcon(QIcon(os.path.join(self.baseDir, "icons", "squish.png")))
        squish.clicked.connect(self.squish)
        vbox.addWidget(squish)

        save = QPushButton("Save")
        save.setIcon(QIcon(os.path.join(self.baseDir, "icons", "save.png")))
        save.clicked.connect(self.save)
        vbox.addWidget(save)

        download = QPushButton("Download")
        download.setIcon(QIcon(os.path.join(self.baseDir, "icons", "download.png")))
        download.clicked.connect(self.download)
        vbox.addWidget(download)

        self.guide = QPushButton("Help")
        self.guide.clicked.connect(self.showGuide)
        self.guide.setIcon(QIcon(os.path.join(self.baseDir, "icons", "help.png")))
        vbox.addWidget(self.guide)

        # comunes = ['Ameglia', 'Arcola', 'Bagnone', 'Bolano']
        # combo = CheckableComboBox()
        # combo.addItems(comunes)
        # combo.toggleItem(2)
        # combo.currentTextChanged.connect(self.textChanged2)
        # vbox.addWidget(combo)

        self.view = QGraphicsView(self.scene)
        self.view.setRenderHint(QPainter.RenderHint.Antialiasing)

        hbox = QHBoxLayout(self)
        hbox.addLayout(vbox)
        hbox.addWidget(self.view)

        self.setLayout(hbox)

        self.darkenFull()
    

    """ FUNCTIONS BELOW """

    def closeEvent(self, event):
        """ Overriden closeEvent to close other windows if still open """
        if self.guideWindow:
            self.guideWindow.close()
        event.accept()
            
    def delete(self):
        """ Delete currently selected blocks """

        for item in self.scene.selectedItems():
            if item.data(2) is not None:
                if item.data(0) is not None and item.data(1) is not None:
                    for cItem in item.collidingItems():
                        if (cItem.data(0) is not None and
                            item.data(0) == cItem.data(0)):
                            for i in range(item.data(1)):
                                self.scene.schedule[item.data(0) + i].isFull = False
                self.scene.removeItem(item)
                self.scene.procedures.remove(item)
                if (item.data(2) != 1):
                    self.currentCaseCount -= 1
                    self.caseCountItem.setPlainText("Cases: " + str(self.currentCaseCount))

        self.darkenFull()

    def clear(self):
        """ Clear all blocks in schedule """

        self.currentCaseCount = 0
        self.caseCountItem.setPlainText("Cases: " + str(self.currentCaseCount))
        self.scene.procedures.clear()

        for item in self.scene.items():
            if item.data(2) is not None:
                self.scene.removeItem(item)
        
        for block in self.scene.schedule:
            block.isFull = False

        self.darkenFull()

    def openFile(self):
        """ Open a xlsx schedule and load into scene """

        fname = QFileDialog.getOpenFileName(self, 'Open file', self.baseDir)
        if fname[0]:
            print(fname[0])
            self.clear()

            # Define variable to load the dataframe
            dataframe = op.load_workbook(fname[0], data_only=True)
            
            # Define variable to read sheet
            dataframe1 = dataframe.active

            softMaxCol = min(12, dataframe1.max_column)
            softMaxRow = min(50, dataframe1.max_row)
            
            colTime = 0
            colFirstName = 0
            colLastName = 0
            colIOL = 0
            colProcedure = 0
            rowFirst = -1
            rowCount = 1
            #TODO: Add more robustness for detecting name & procedure
            # Find column indices for desired data
            for row in dataframe1.iter_rows(min_row=1, max_col=4, max_row=30):
                for cell in row:
                    if cell.value is not None:
                        val = cell.value.lower()
                        if "time" in val:
                            rowFirst = rowCount + 1
                            break
                if rowFirst != -1:
                    break
                else:
                    rowCount += 1
            
            for row in dataframe1.iter_rows(rowCount, rowCount):
                count = 0
                for cell in row:
                    if cell.value is not None:
                        val = cell.value.lower()
                        
                        if "first" in val:
                            colFirstName = count
                        elif "last" in val:
                            colLastName = count
                        elif "primary iol" in val:
                            colIOL = count
                        elif "special notes" in val:
                            colProcedure = count
                    count += 1
       
            try:
                # Iterate through rows and add procedures to schedule

                for row in dataframe1.iter_rows(rowFirst, softMaxRow):
                    stopLoad = False
                    diffBreak = False
                    if  row[colFirstName].value is None and row[colLastName].value is None and row[colIOL].value is None:
                        for col in row:
                            if col.value is not None:
                                try:
                                    if "cancelled" in col.value.lower():
                                        stopLoad = True
                                        break
                                    elif "Surgeon Break" in col.value:
                                        stopLoad = False
                                        diffBreak = True
                                        break
                                    else:
                                        stopLoad = True     
                                except Exception as e:
                                    print("Could not read column:", e)

                            if diffBreak:
                                break
          
                    if stopLoad:
                        break
                    else:
                        firstEmpty = -1
                        firstY = -1
                        blockTypes = []
                        blockType = "Regular"
                        customLength = 1

                        if (diffBreak or
                            row[colIOL].value is not None and "Surgeon Break" in row[colIOL].value or
                            row[colFirstName].value is not None and "Surgeon Break" in row[colFirstName].value or
                            row[colLastName].value is not None and "Surgeon Break" in row[colLastName].value):
                            blockType = "Break"
                            blockTypes.append("Break")
                            name = "Break"
                        else:
                            name = row[colFirstName].value[0] + ". " + row[colLastName].value
                            procedure = row[colProcedure].value.lower()
                            
                            # check for block type and add it
                            if ("toric" in procedure):
                                blockTypes.append("Premium")
                            elif ("vivity" in procedure):
                                blockTypes.append("Vivity")
                            elif ("panoptix" in procedure):
                                blockTypes.append("PanOptix")
                            elif ("femto" in procedure or "lensx" in procedure or "/ora" in procedure or "catalyst" in procedure):
                                blockTypes.append("Laser")
                            elif "phaco" in procedure or "trimox" in procedure:
                                blockTypes.append("Regular")

                            if ("xen" in procedure):
                                if ("revision" in procedure):
                                    blockTypes.append("XENRevision")
                                else:
                                    blockTypes.append("XEN")
                            elif ("wound" in procedure and "revision" in procedure):
                                blockTypes.append("WoundRevision")

                            if ("goniotomy" in procedure or "goniosynechialysis" in procedure):
                                #blockTypes.append("Goniotomy")
                                blockTypes.append("MIGS")
                            if ("stent" in procedure):
                                #blockTypes.append("Stent")
                                blockTypes.append("MIGS")
                            if ("canaloplasty" in procedure):
                                #blockTypes.append("Canaloplasty")
                                blockTypes.append("MIGS")
                            if ("shunt" in procedure):
                                blockTypes.append("Shunt")
                            if ("trabeculectomy" in procedure):
                                blockTypes.append("Trabeculectomy")
                            if ("micropulse" in procedure):
                                blockTypes.append("Micropulse")
                            if (len(blockTypes) == 0):
                                blockTypes.append("Custom")

                        #     if ("shunt" in procedure):
                        #         blockType = "Shunt"
                        #     elif ("xen" in procedure and "revision" not in procedure):
                        #         blockType = "XEN"
                        #     elif ("trabeculectomy" in procedure):
                        #         blockType = "Trabeculectomy"
                        #     elif ("xen" in procedure and "revision" in procedure):
                        #         blockType = "XENRevision"
                        #     elif ("wound" in procedure and "revision" in procedure):
                        #         blockType = "WoundRevision"
                        #     elif ("vivity" in procedure):
                        #         blockType = "Vivity"
                        #     elif ("panoptix" in procedure):
                        #         blockType = "PanOptix"
                        #     elif ("toric" in procedure):
                        #         blockType = "Premium"
                        #     elif ("goniotomy" in procedure or "goniosynechialysis" in procedure):
                        #         blockType = "Goniotomy"
                        #     elif ("femto" in procedure or "lensx" in procedure or "/ora" in procedure or "catalyst" in procedure):
                        #         blockType = "Laser"
                        #     elif ("micropulse" in procedure):
                        #         blockType = "Micropulse"
                        #     elif ("canaloplasty" in procedure):
                        #         blockType = "Canaloplasty"
                        #     elif ("stent" in procedure):
                        #         blockType = "Stent"
                        #     elif "phaco" in procedure or "trimox" in procedure:
                        #         blockType = "Regular"
                        #     else:
                        #         blockType = "Custom"

                        # if blockType != "Custom":
                        #     segLength = int(self.blockTimes[blockType] * self.numSubBlocks)
                        #     length = self.blockTimes[blockType]
                        # else:
                        #     segLength = int(customLength * self.numSubBlocks)
                        #     length = customLength

                        if blockTypes[0] != "Custom":
                            segLength = 0
                            length = 0
                            for type in blockTypes:
                                length += self.blockTimes[type]
                            segLength = int(length * self.numSubBlocks)
                        else:
                            segLength = int(customLength * self.numSubBlocks)
                            length = customLength
                        
                        
                        firstEmpty, firstY = self.findFirstEmpty(segLength)
           
                        if (firstEmpty == -1):
                            continue
                        else:
                            print(blockTypes)
                            #self.insertBlock(blockType, firstEmpty, firstY, segLength, length, name)
                            self.insertBlock(blockTypes, firstEmpty, firstY, segLength, length, name)
    
                self.caseCountItem.setPlainText("Cases: " + str(self.currentCaseCount))
                self.darkenFull()
                        

            except Exception as e:
                print("Could not read schedule", e)
    
    def squish(self):
        """ Using the Graphics Scene squish method, remove whitespace between blocks """
        self.scene.squish(0, len(self.scene.schedule) -1 )
        self.darkenFull()
        return

    def save(self):
        """ Save current schedule to SavedSchedule.txt """

        try:
            with open(os.path.join(self.baseDir, "resources", "SavedSchedule.txt"), "w") as scheduleFile:
                schedule = np.zeros((len(self.scene.schedule), 2))
                scheduleNames = [""] * len(self.scene.schedule)
                scheduleStr = ''
                for item in self.scene.items():
                    if item.data(2) is not None:
                        schedule[item.data(0), 0] = item.data(2)
                        for i in range(item.data(1)):
                            schedule[item.data(0) + i, 1] = 1

                        subItems = item.childItems()
                        for subItem in subItems:
                            if subItem.data(3) is not None and subItem.data(3) == 1:
                                #scheduleNames[item.data(0)] = "," + subItem.toPlainText() + "," + str(item.data(1))
                                scheduleNames[item.data(0)] = "," + subItem.toPlainText() + "," + str(item.data(1)) + "," + item.blockStr

                for i in range(len(schedule)):
                    if schedule[i][0] == 0 and schedule[i][1] != 1:
                        scheduleStr += "0\n"
                    elif schedule[i][0] != 0:
                        scheduleStr += (str(int(schedule[i][0])) + scheduleNames[i] + "\n")

                scheduleFile.write(scheduleStr)
                print("Saved schedule")
        except Exception as e:
            print("Could not save schedule:", e)

    def download(self):
        """ Download current schedule as a JPG image """
        try:
            pix = QPixmap(self.canvasWidth, self.canvasHeight)
            painter = QPainter(pix)
            self.scene.render(painter)
            painter.end()
            pix.save("schedule-" + str(date.today()) + ".jpg", "JPG")
        except Exception as e:
            print("Could not save:", e)

    def darkenFull(self):
        """ Utility function to visualize which segments are currently full """

        # print("Procedures: ", len(self.scene.procedures))

        for item in self.scene.items():
            if item.data(4) is not None:
                if self.scene.schedule[item.data(0)].isFull == True:
                    item.setBrush(QColor(0, 0, 0, 125))
                    item.update()
                else: 
                    item.setBrush(QColor(0, 0, 0, 0))
                    item.update()
        return

    
    def insertSaved(self, inputType = -1, skip = 0, name = "Patient", segLength = -1, blockTypes = [0]):
        """" Insert a new schedule block for saved schedule"""
        
        if inputType == 0:
            return
        else:
            #blockType = ProcedureDicts.procedureList[inputType]
            blockType = blockTypes[0]
        # Find first empty block in scene that can accommodate new insertion
        if blockType == "Custom":
            length = segLength / (1.0 * self.numSubBlocks)
        else:
            length = 0.0
            for type in blockTypes:
                length += self.blockTimes[type]
        
        firstEmpty = -1
        firstY = -1
        if (skip < len(self.scene.schedule) and skip + segLength - 1 < len(self.scene.schedule)):
            firstEmpty = skip
            firstY = self.scene.schedule[firstEmpty].y
        
        if firstEmpty != -1:
            self.insertBlock(blockTypes, firstEmpty, firstY, segLength, length, name)

    def insert(self):
        """" Insert a new schedule block at the first accommodating empty slot """
        
        blockType = self.blockType
        
        segLength = int(self.blockTimes[blockType] * self.numSubBlocks)
        
        firstEmpty, firstY = self.findFirstEmpty(segLength)

        if firstEmpty != -1:

            if blockType == "Break":
                rectText = "Break"
            else:
                rectText = self.blockName
            self.insertBlock([blockType], firstEmpty, firstY, segLength, self.blockTimes[blockType], rectText)
        self.darkenFull()
    
    def insertBlock(self, blockTypes, firstEmpty, firstY, segLength, length, rectText):
        """ Create and add a procedure block at the given empty index to the scene """
    
        timeText = self.getTimeText(firstEmpty, firstEmpty + segLength - 1)
        rect = GraphicsRectItem(0, 0, self.blockWidth, self.blockSize * length, self.timeWidth, self.typeWidth, 
                                timeText, rectText, blockTypes, self.blockColors[blockTypes[0]], segLength)
        rect.setPos(self.timeBoxWidth, firstY)
        rect.setData(0, firstEmpty)                            # id of first block segment that rect occupies
        rect.setData(1, segLength)                             # length that of procedure
        rect.setData(2, ProcedureDicts.procedureIDs[blockTypes[0]])   # identifier for graphics to tell that this is a rect

        self.scene.procedures.append(rect)
        self.scene.addItem(rect)
        self.fill(firstEmpty, segLength)

        rect.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable)
        rect.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable)

        if (ProcedureDicts.procedureIDs[blockTypes[0]] != 1):
            self.currentCaseCount += 1
            self.caseCountItem.setPlainText("Cases: " + str(self.currentCaseCount))

    def getTimeText(self, startIndex, endIndex):
        """ Return a string that states the time range from the startIndex to endIndex """
        try:
            return self.scene.schedule[startIndex].beginString + " - " + self.scene.schedule[endIndex].endString
        except Exception as e:
            print("There was an issue populating break time:", e)
            return "Time"
    
    def fill(self, startIndex, segLength, isBreak = False):
        """ Set the schedule fullness according to parameters """
        try:
            for i in range(segLength):
                self.scene.schedule[startIndex + i].isFull = True
        except Exception as e:
            print("There was an issue:", e)

    def findFirstEmpty(self, segLength, startIndex = 0, endIndex = 1000000):
        """ Find the first empty section from the startIndex to endIndex (incl.) of length segLength """
        endIndex = min(len(self.scene.schedule), endIndex + 1)
        if (startIndex + segLength < endIndex):
            for i in range(startIndex, endIndex):
                try:
                    if (self.scene.schedule[i].isFull == False):
                        isColliding = False
                        
                        for j in range(1, segLength):
                            if (self.scene.schedule[i + j].isFull == True):
                                isColliding = True
                                i = i + j + 1
                                break

                        if (not isColliding):
                            return self.scene.schedule[i].order, self.scene.schedule[i].y
                        
                except Exception as e:
                    print("Issue checking collisions:", e)
                    break
        return -1, -1

    def index_changed(self, i): # i is an int
        print(i)
    
    def textChanged2(self, s):
        for item in self.combo.currentData():
            print(item)
        #print(s)
        print("END")

    def text_changed(self, s): # s is a str
        """ Change the stored procedure type when combobox is updated """

        if (s == "Custom" and not self.customLength.isEnabled()):
            self.customLength.setEnabled(True)
            self.customLengthLabel.setEnabled(True)
        elif (s != "Custom" and self.customLength.isEnabled()):
            self.customLength.setEnabled(False)
            self.customLengthLabel.setEnabled(False)

        self.blockType = s

    def showGuide(self):
        if self.guideWindow is None:
            self.guideWindow = HelpGuide()
        self.guideWindow.show()
    
    def changeCustomLength(self, i):
        """ Change the stored custom length type when the appropriate widget is updated """
        self.blockTimes["Custom"] = i
        self.customBlockLength = i

    def name_changed(self, s):
        """ Change the stored patient name to insert when the text box is updated """
        self.blockName = s
    
    def nameModified(self, s):
        """ Change the stored patient name to modify when the text box is updated """
        self.blockModName = s
    
    def changePatientName(self):
        """ Change the name of the selected block """
        self.scene.changeName(self.blockModName)

    def changeInputPatientName(self, name="Patient"):
        """ Change the name of the text input widget for insertion and modification """
        self.patientName.setText(name)
        self.modifyPatientName.setText(name)

    def enableNameChange(self, enable=False):
        """ Enable/Disable the change name and delete buttons on widget """
        self.changeName.setEnabled(enable)
        self.deleteBlock.setEnabled(enable)
